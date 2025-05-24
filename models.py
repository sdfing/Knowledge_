from sklearn.preprocessing import OneHotEncoder, LabelEncoder
from werkzeug.security import generate_password_hash, check_password_hash
from database import db
from sqlalchemy import text
from sqlalchemy import ForeignKey
from sqlalchemy.orm import relationship
from sklearn.linear_model import LinearRegression
from sklearn.ensemble import GradientBoostingRegressor
import numpy as np


class User(db.Model):
    __tablename__ = 'users'
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(64), unique=True, index=True)
    password_hash = db.Column(db.String(128))

    def set_password(self, password):
        self.password_hash = generate_password_hash(password)

    def check_password(self, password):
        return check_password_hash(self.password_hash, password)


class Course(db.Model):
    __tablename__ = 'course'
    id = db.Column(db.Integer, primary_key=True, name='课程代码')
    name = db.Column(db.String(64), unique=True, nullable=False, name='课程名称')

    # ... other course fields ...

    @staticmethod
    def extract_and_save_course_info():
        """
        Extract all courses from the enrollment table and save them to the course table.
        """
        # Query the enrollment table for all distinct courses
        sql_query = text("""
            SELECT DISTINCT
                课程代码,
                课程名称
            FROM
                enrollment
        """)
        results = db.session.execute(sql_query)

        # Iterate through the results and create or update course records
        for row in results:
            course_code, course_name = row
            # Check if the course already exists
            existing_course = Course.query.filter_by(id=course_code).first()
            if not existing_course:
                # Course does not exist, create a new record
                new_course = Course(
                    id=course_code,
                    name=course_name
                )
                db.session.add(new_course)

        # Commit all new records to the database
        db.session.commit()

    def __repr__(self):
        return '<Course %r>' % self.name


class Student(db.Model):
    __tablename__ = 'student'  # 学生表名
    student_id = db.Column(db.String(255), primary_key=True, name='学号')
    name = db.Column(db.String(255), name='姓名')
    password_hash = db.Column(db.String(255), name='登录密码')
    major = db.Column(db.String(255), name='专业')
    class_id = db.Column(db.String(255), name='班级')
    gender = db.Column(db.String(255), name='性别')

    # knowledge_scores = relationship('StudentKnowledge', back_populates='student')

    # ... 其他个人信息字段 ...

    def set_password(self, password):
        self.password_hash = generate_password_hash(password)

    @staticmethod
    def get_or_create_student(student_id, name, major, class_id, password):

        """
        从选课表中查询学生的信息，如果学生不存在则创建新记录。

        :param student_id: 学生的学号。
        :param name: 学生的姓名。
        :param major: 学生的专业。
        :param class_id: 学生的班级。
        :param password: 学生的登录密码。
        :return: Student对象。
        """
        student = Student.query.filter_by(student_id=student_id).first()
        if not student:
            student = Student(
                student_id=student_id,
                name=name,
                major=major,
                class_id=class_id,
                password_hash='1'
            )
            student.set_password(password)
            db.session.add(student)
            db.session.commit()
        return student

    @staticmethod
    def extract_and_save_student_info():
        """
        从选课表中提取所有学生的信息，并保存到学生信息表中。
        """
        # 查询选课表中的所有不同学生的信息
        sql_query = text("""
            SELECT DISTINCT
                学号,
                姓名,
                专业,
                班级,
                性别
            FROM
                enrollment
        """)
        results = db.session.execute(sql_query)

        # 遍历查询结果，为每个学生创建或更新记录
        for row in results:
            student_id, name, major, class_id, gender = row
            # 检查学生是否已存在
            existing_student = Student.query.filter_by(student_id=student_id).first()
            if not existing_student:
                # 学生不存在，创建新记录
                new_student = Student(
                    student_id=student_id,
                    name=name,
                    major=major,
                    class_id=class_id,
                    gender=gender
                )
                # 设置一个默认密码，实际应用中应该让用户自己设置或生成随机密码
                default_password = 'default_password'
                # 设置默认密码或生成随机密码
                new_student.set_password(default_password)
                db.session.add(new_student)

        # 提交所有新记录到数据库
        db.session.commit()


class KnowledgePoint(db.Model):
    __tablename__ = 'knowledge_point'
    KnowledgeID = db.Column(db.Integer, name='知识点id', primary_key=True)
    KnowledgeName = db.Column(db.String(255), name='知识点名称', unique=True)
    Chapter = db.Column(db.String, name='章节', primary_key=True)
    KnowledgeBelong = db.Column(db.String(255), name='所属课程')

    @staticmethod
    def update_scoring_from_excel(file_path):
        import openpyxl
        # 加载Excel工作簿
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active

        # 获取标题行
        header = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))

        # 查找包含"知识点"关键词的列索引
        knowledge_point_indexes = [i for i, col in enumerate(header) if '知识点' in col]

        # 遍历Excel表格中的每一行（跳过标题行）
        for row in sheet.iter_rows(min_row=2, values_only=True):
            scoring = row[0]  # 获取计分环节

            # 遍历包含"知识点"关键词的列
            for index in knowledge_point_indexes:
                knowledge_point = row[index]
                if knowledge_point:
                    # 根据知识点名称查询数据库中对应的记录
                    knowledge = KnowledgePoint.query.filter_by(KnowledgeName=knowledge_point).first()
                    if knowledge:
                        # 如果找到对应的记录，更新计分环节字段
                        knowledge.Scoring = scoring
                        db.session.commit()

        print("计分环节更新完成")


class StagePreYear(db.Model):
    __tablename__ = 'stage_pre_year'
    KnowledgeID = db.Column(db.Integer, name='知识点id', primary_key=True)
    Year = db.Column(db.String(100), name='年份', primary_key=True)
    Scoring = db.Column(db.String(100), name='计分环节')
    Stage = db.Column(db.String(100), name='阶段')

    @staticmethod
    def update_stage_pre_year_from_excel(file_path, year):
        import openpyxl
        # 加载Excel工作簿
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active

        # 获取标题行
        header = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))

        # 查找包含"知识点"关键词的列索引
        knowledge_point_indexes = [i for i, col in enumerate(header) if '知识点' in col]

        # 创建一个字典来记录每个知识点已经记录过的计分环节
        knowledge_scoring_dict = {}

        # 遍历Excel表格中的每一行(跳过标题行)
        for row in sheet.iter_rows(min_row=2, values_only=True):
            scoring = row[0]  # 获取计分环节

            # 遍历包含"知识点"关键词的列
            for index in knowledge_point_indexes:
                knowledge_point = row[index]
                if knowledge_point:
                    # 根据知识点名称查询数据库中对应的知识点ID
                    knowledge = KnowledgePoint.query.filter_by(KnowledgeName=knowledge_point).first()
                    if knowledge:
                        knowledge_id = knowledge.KnowledgeID

                        # 检查该知识点是否已经记录过计分环节
                        if knowledge_id in knowledge_scoring_dict:
                            continue  # 如果已经记录过,则跳过后续操作

                        # 记录该知识点的计分环节
                        knowledge_scoring_dict[knowledge_id] = scoring

                        # 查询 StagePreYear 表中是否已存在对应的记录
                        stage_pre_year = StagePreYear.query.filter_by(KnowledgeID=knowledge_id, Year=year).first()
                        if stage_pre_year:
                            # 如果记录已存在,更新计分环节和阶段字段
                            stage_pre_year.Scoring = scoring
                            stage_pre_year.Stage = StagePreYear.stage_by_scoring(scoring)  # 根据实际情况设置阶段值
                        else:
                            # 如果记录不存在,创建新记录并添加到数据库
                            new_stage_pre_year = StagePreYear(KnowledgeID=knowledge_id, Year=year, Scoring=scoring,
                                                              Stage=StagePreYear.stage_by_scoring(scoring))
                            db.session.add(new_stage_pre_year)
                        db.session.commit()

        print(f"{year}年预期阶段更新完成")

    @staticmethod
    def stage_by_scoring(scoring):
        stage1_scorings = ['作业1线性', '实验1', '实验2']
        stage2_scorings = ['作业2树图', '实验3', '实验4']
        stage3_scorings = ['阶段考试','作业3查找排序']
        stage4_scorings = ['期末考试']

        if scoring in stage1_scorings:
            return "阶段1线性"
        elif scoring in stage2_scorings:
            return "阶段2树图"
        elif scoring in stage3_scorings:
            return "阶段3查找排序"
        elif scoring in stage4_scorings:
            return "阶段4期末"
        else:
            return None


class KnowledgePointEdge(db.Model):
    __tablename__ = 'knowledge_point_edge'
    sourceID = db.Column(db.Integer, name='source', primary_key=True)
    targetID = db.Column(db.Integer, name='target', primary_key=True)
    KnowledgeBelong = db.Column(db.String(255), name='所属课程')
    # 其他字段...


class Enrollment(db.Model):
    __tablename__ = 'enrollment'  # 表名选课表
    class_id = db.Column(db.String(255), name='班级')
    student_id = db.Column(db.String(255), primary_key=True, name='学号')
    name = db.Column(db.String(255), name='姓名')
    course_name = db.Column(db.String(255), name='课程名称')
    grade = db.Column(db.String(255), name='成绩')
    academic_year = db.Column(db.String(255), name='学年')
    semester = db.Column(db.String(255), name='学期')
    student_category = db.Column(db.String(255), name='学生类别')
    college = db.Column(db.String(255), name='学院')
    major = db.Column(db.String(255), name='专业')
    major_direction = db.Column(db.String(255), name='专业方向', nullable=True)
    grade_year = db.Column(db.String(255), name='年级', nullable=True)
    student_tag = db.Column(db.String(255), name='学生标记', nullable=True)
    course_college = db.Column(db.String(255), name='开课学院')
    course_code = db.Column(db.String(255), name='课程代码')
    teaching_class = db.Column(db.String(255), name='教学班')
    instructor = db.Column(db.String(255), name='任课教师')
    credit = db.Column(db.Float, name='学分')
    grade_remark = db.Column(db.String(255), name='成绩备注', nullable=True)
    exam_nature = db.Column(db.String(255), name='考试性质')
    gpa = db.Column(db.Float, name='绩点')
    course_tag = db.Column(db.String(255), name='课程标记', nullable=True)
    course_category = db.Column(db.String(255), name='课程类别')
    course_belonging = db.Column(db.String(255), name='课程归属')
    course_nature = db.Column(db.String(255), name='课程性质')
    assessment_method = db.Column(db.String(255), name='考核方式')
    is_grade_invalid = db.Column(db.String(255), name='是否成绩作废')
    submitter = db.Column(db.String(255), name='提交人')
    submission_time = db.Column(db.DateTime, name='提交时间')
    is_degree_course = db.Column(db.String(255), name='是否学位课程')
    single_class_retake = db.Column(db.String(255), name='单开班重修', nullable=True)
    gender = db.Column(db.String(255), name='性别')
    remark_info = db.Column(db.String(255), name='备注信息', nullable=True)
    credit_point = db.Column(db.Float, name='学分绩点', nullable=True)
    course_type = db.Column(db.String(255), name='开课类型')

    def calculate_credit_points(student_id, excel_file):
        from sqlalchemy import text
        from openpyxl import load_workbook

        # 定义成绩对应的分数
        grade_to_score = {
            '优秀': 95,
            '良好': 85,
            '中等': 75,
            '及格': 65
        }

        # 读取Excel表格
        wb = load_workbook(excel_file)
        sheet = wb.active

        # 获取表格中的课程名称列表
        course_names_in_excel = [row[0] for row in sheet.iter_rows(min_row=2, values_only=True)]

        # 查询数据库获取学生选修的课程、学分绩点和成绩
        sql_query = text("""
            SELECT 
                课程名称,
                学分绩点,
                成绩
            FROM 
                enrollment
            WHERE 
                学号 = :student_id
        """)
        # 将查询结果转换为列表，以便可以多次迭代
        db_results = list(db.session.execute(sql_query, {'student_id': student_id}))

        total_credit_points = 0
        selected_courses_in_excel = []
        missing_courses_in_db = []
        course_grades = []
        courses_with_18_credit_points = []  # 用于存储学分绩为18的课程

        # 获取学生在数据库中选修的课程名称和学分绩列表
        course_dict_in_db = {row[0]: {'credit_point': row[1], 'grade': row[2]} for row in db_results}

        # 找出学生选修的且在表格中的课程
        for course_name in course_dict_in_db:
            if course_name in course_names_in_excel:
                selected_courses_in_excel.append(course_name)

        # 找出在表格中但学生未选修的课程
        for course_name in course_names_in_excel:
            if course_name not in course_dict_in_db:
                missing_courses_in_db.append(course_name)

        # 计算学生选修的且在表格中的课程的学分绩之和，并记录成绩
        for course_name in selected_courses_in_excel:
            credit_point = course_dict_in_db[course_name]['credit_point']
            grade = course_dict_in_db[course_name]['grade']

            # 如果成绩是文字（如“优秀”），则转换为相应的分数，否则直接使用数值成绩
            if grade in grade_to_score:
                score = grade_to_score[grade]
            else:
                score = float(grade)  # 假设其他情况都是数值成绩，并转换为浮点数

            total_credit_points += credit_point if credit_point else 0
            course_grades.append({
                'course_name': course_name,
                'grade': grade,
                'score': score
            })

            # 检查当前课程的学分绩是否为18，如果是则记录下来
            if credit_point == 18:
                courses_with_18_credit_points.append(course_name)

        # 按照成绩（分数）从低到高排序
        course_grades.sort(key=lambda x: x['score'])

        # 打印结果
        print(f"学号为 {student_id} 的学生选修了以下在表格中的课程：")
        for course in course_grades:
            print(
                f"{course['course_name']} - 学分绩: {course_dict_in_db[course['course_name']]['credit_point']} - 成绩: {course['grade']} ({course['score']}分)")

        print(f"\n学分绩之和：{total_credit_points}")

        if missing_courses_in_db:
            print("\n学生未选修的表格中的课程：")
            for course_name in missing_courses_in_db:
                print(course_name)
        else:
            print("\n学生选修了表格中的所有课程。")

        # 打印学分绩为18的课程
        if courses_with_18_credit_points:
            print("\n学分绩为18的课程：")
            for course_name in courses_with_18_credit_points:
                print(course_name)
        else:
            print("\n没有学分绩为18的课程。")



    @staticmethod
    def rank_avg_gpa(major, grade):
        sql_query = text("""
            SELECT 
                学号, 
                姓名, 
                专业, 
                ROUND(SUM(学分绩点) / SUM(学分), 2) AS 平均学分绩点
            FROM 
                enrollment
            WHERE 
                年级 = :grade AND
                专业 = :major
            GROUP BY 
                学号, 姓名, 专业
        """)
        db_results = db.session.execute(sql_query, {'major': major, 'grade': grade})
        results = list(db_results)

        for i in range(len(results)):
            results[i] = list(results[i])
            name = results[i][1]
            encrypted_name = name[0] + '*' * (len(name) - 1)
            results[i][1] = encrypted_name

        results.sort(key=lambda x: x[3], reverse=True)

        for i in range(len(results)):
            results[i].append(i + 1)

        return results


    @staticmethod
    def map_score(score):
        if score == '优':
            return 95
        elif score == '良':
            return 85
        elif score == '中':
            return 75
        elif score == '及格':
            return 65
        elif score == '不及格' or score == '缓考':
            return 0
        else:
            try:
                score = float(score)
                return score
            except ValueError:
                return None

    @staticmethod
    def get_violin_data_roughly(grade_year, course_name_search, major, semester, academic_year, is_exam_course,
                                exam_nature):
        grade_year = '%' + grade_year + '%'
        course_name_search = '%' + course_name_search + '%'
        major = '%' + major + '%'
        semester = '%' + semester + '%'
        academic_year = '%' + academic_year + '%'

        sql_query = text("""
               SELECT 
                   班级,
                   成绩,
                   课程名称
               FROM 
                   enrollment
               WHERE 
                   年级 LIKE :grade_year AND
                   课程名称 LIKE :course_name_search AND
                   专业 LIKE :major AND
                   学期 LIKE :semester AND
                   学年 LIKE :academic_year AND
                   考试性质 LIKE :exam_nature
           """)
        db_results = db.session.execute(sql_query, {
            'grade_year': grade_year,
            'course_name_search': course_name_search,
            'major': major,
            'semester': semester,
            'academic_year': academic_year,
            'exam_nature': exam_nature,
        })

        violin_data = []
        for class_id, grade, course_name in db_results:
            if is_exam_course == '1':
                try:
                    grade = float(grade)
                    violin_data.append({
                        "courseName": course_name,
                        "x": class_id,
                        "y": grade
                    })
                except ValueError:
                    # print('passed')
                    pass
            else:
                mapped_score = Enrollment.map_score(grade)
                if mapped_score is not None:
                    violin_data.append({
                        "courseName": course_name,
                        "x": class_id,
                        "y": mapped_score
                    })

        return violin_data

    @staticmethod
    def get_self_home_data(semester, academic_year, student_id):
        semester = '%' + semester + '%'
        academic_year = '%' + academic_year + '%'
        student_id = student_id

        sql_query = text("""
            SELECT 
                e.课程名称,
                CASE 
                    WHEN e.成绩 = '优秀' THEN 95
                    WHEN e.成绩 = '良好' THEN 85
                    WHEN e.成绩 = '中等' THEN 75
                    WHEN e.成绩 = '及格' THEN 60
                    WHEN e.成绩 = '通过' THEN 100
                    WHEN e.成绩 = '不及格' THEN 0
                    ELSE CAST(e.成绩 AS DECIMAL)
                END AS 个人成绩,
                AVG(
                    CASE 
                        WHEN e2.成绩 = '优秀' THEN 95
                        WHEN e2.成绩 = '良好' THEN 85
                        WHEN e2.成绩 = '中等' THEN 75
                        WHEN e2.成绩 = '及格' THEN 60
                        WHEN e2.成绩 = '通过' THEN 100
                        WHEN e2.成绩 = '不及格' THEN 0
                        ELSE CAST(e2.成绩 AS DECIMAL)
                    END
                ) AS 班级平均成绩,
                AVG(
                    CASE 
                        WHEN e3.成绩 = '优秀' THEN 95
                        WHEN e3.成绩 = '良好' THEN 85
                        WHEN e3.成绩 = '中等' THEN 75
                        WHEN e3.成绩 = '及格' THEN 60
                        WHEN e3.成绩 = '通过' THEN 100
                        WHEN e3.成绩 = '不及格' THEN 0
                        ELSE CAST(e3.成绩 AS DECIMAL)
                    END
                ) AS 专业平均成绩
            FROM 
                enrollment e
            LEFT JOIN 
                enrollment e2 ON e.课程名称 = e2.课程名称 AND e.班级 = e2.班级
            LEFT JOIN
                enrollment e3 ON e.课程名称 = e3.课程名称 AND e.专业 = e3.专业
            WHERE 
                e.学期 LIKE :semester AND
                e.学年 LIKE :academic_year AND
                e.学号 LIKE :student_id
            GROUP BY
                e.课程名称, e.成绩
        """)
        db_results = db.session.execute(sql_query, {
            'semester': semester,
            'academic_year': academic_year,
            'student_id': student_id,
        })

        radar_data = []
        line_data = []
        for subject, personal_score, class_avg_score, major_avg_score in db_results:
            # 处理过长学科名称
            if (len(subject) > 8):
                subject = subject[:4] + '..' + subject[-2:]
            if personal_score is not None:
                radar_data.append({
                    "subject": subject,
                    "type": "个人",
                    "score": float(personal_score)
                })
                line_data.append({
                    "subject": subject,
                    "个人分数": float(personal_score),
                    "班级平均分": float(class_avg_score) if class_avg_score else None,
                    "专业平均分": float(major_avg_score) if major_avg_score else None
                })
            if class_avg_score is not None:
                radar_data.append({
                    "subject": subject,
                    "type": "班级平均",
                    "score": float(class_avg_score)
                })

        return {"radarData": radar_data, "lineData": line_data}

    @staticmethod
    def get_violin_data_value_added(grade_year, course_name_search, major, semester, academic_year, is_exam_course,
                                    exam_nature):
        grade_year = '%' + grade_year + '%'
        course_name_search = '%' + course_name_search + '%'
        major = '%' + major + '%'
        semester = '%' + semester + '%'
        academic_year = '%' + academic_year + '%'

        sql_query = text("""
               SELECT 
                   班级,
                   CASE 
                       WHEN 成绩 = '优秀' THEN 95
                       WHEN 成绩 = '良好' THEN 85
                       WHEN 成绩 = '中等' THEN 75
                       WHEN 成绩 = '及格' THEN 60
                       WHEN 成绩 = '通过' THEN 100
                       WHEN 成绩 = '不及格' THEN 0
                       ELSE CAST(成绩 AS DECIMAL)
                   END AS gpa,
                   课程名称,
                   学年,
                   学期
               FROM 
                   enrollment
               WHERE 
                   年级 LIKE :grade_year AND
                   课程名称 LIKE :course_name_search AND
                   专业 LIKE :major AND
                   学期 LIKE :semester AND
                   学年 LIKE :academic_year AND
                   考试性质 LIKE :exam_nature
               ORDER BY
                   班级, 学年, 学期
           """)
        db_results = db.session.execute(sql_query, {
            'grade_year': grade_year,
            'course_name_search': course_name_search,
            'major': major,
            'semester': semester,
            'academic_year': academic_year,
            'exam_nature': exam_nature,
        })

        # 将数据按照班级分组
        class_data = {}
        for class_id, gpa, course_name, academic_year, semester in db_results:
            if class_id not in class_data:
                class_data[class_id] = []
            class_data[class_id].append({
                'gpa': gpa,
                'course_name': course_name,
                'academic_year': academic_year,
                'semester': semester
            })

        # 计算每个班级的增值评价
        violin_data = []
        for class_id, enrollments in class_data.items():
            if len(enrollments) > 1:
                # 提取学期和GPA数据
                semesters = []
                gpas = []
                for i, enrollment in enumerate(enrollments):
                    semesters.append(i)
                    gpas.append(enrollment['gpa'])

                # 使用线性回归模型进行增值评价
                model = LinearRegression()
                model.fit([[semester] for semester in semesters], gpas)
                value_added = model.coef_[0]

                violin_data.append({
                    "courseName": enrollments[0]['course_name'],
                    "x": class_id,
                    "y": value_added
                })

        return violin_data

    # @staticmethod
    # def get_self_home_data_value_added(student_id):
    #     sql_query = text("""
    #         SELECT
    #             e.学年,
    #             e.学期,
    #             s.班级,
    #             s.专业,
    #             SUM(e.学分绩点) AS 学期学分绩点总和,
    #             SUM(e.学分) AS 学期学分总和,
    #             COALESCE(c.班级学分绩点总和 / c.班级学分总和, 0) AS 班级平均学期绩点,
    #             COALESCE(m.专业学分绩点总和 / m.专业学分总和, 0) AS 专业平均学期绩点
    #         FROM
    #             enrollment e
    #             JOIN student s ON e.学号 = s.学号
    #             LEFT JOIN (
    #                 SELECT
    #                     e2.学年,
    #                     e2.学期,
    #                     s2.班级,
    #                     SUM(e2.学分绩点) AS 班级学分绩点总和,
    #                     SUM(e2.学分) AS 班级学分总和
    #                 FROM
    #                     enrollment e2
    #                     JOIN student s2 ON e2.学号 = s2.学号
    #                 GROUP BY
    #                     e2.学年, e2.学期, s2.班级
    #             ) c ON e.学年 = c.学年 AND e.学期 = c.学期 AND s.班级 = c.班级
    #             LEFT JOIN (
    #                 SELECT
    #                     e3.学年,
    #                     e3.学期,
    #                     s3.专业,
    #                     SUM(e3.学分绩点) AS 专业学分绩点总和,
    #                     SUM(e3.学分) AS 专业学分总和
    #                 FROM
    #                     enrollment e3
    #                     JOIN student s3 ON e3.学号 = s3.学号
    #                 GROUP BY
    #                     e3.学年, e3.学期, s3.专业
    #             ) m ON e.学年 = m.学年 AND e.学期 = m.学期 AND s.专业 = m.专业
    #         WHERE
    #             e.学号 LIKE :student_id
    #         GROUP BY
    #             e.学年, e.学期, s.班级, s.专业
    #         ORDER BY
    #             e.学年, e.学期
    #     """)
    #     db_results = db.session.execute(sql_query, {
    #         'student_id': student_id,
    #     })
    #
    #     # 提取学期、班级、专业和学期平均分数据
    #     semesters = []
    #     classes = []
    #     majors = []
    #     semester_average_scores = []
    #     class_average_scores = []
    #     major_average_scores = []
    #     for (academic_year, semester, class_name, major, total_credit_score, total_credit, class_average_score,
    #          major_average_score) in db_results:
    #         semesters.append(f"{academic_year} {semester}")
    #         classes.append(class_name)
    #         majors.append(major)
    #         semester_average_scores.append(float(total_credit_score) / float(total_credit))
    #         class_average_scores.append(float(class_average_score) if class_average_score else 0)
    #         major_average_scores.append(float(major_average_score) if major_average_score else 0)
    #
    #     print(semester_average_scores)
    #     print(class_average_scores)
    #     print(major_average_scores)
    #     print(semesters)
    #
    #     # 建立个人线性回归模型并计算每个学期的残差
    #     personal_model = LinearRegression()
    #     personal_model.fit([[i] for i in range(len(semesters))], semester_average_scores)
    #     personal_residuals = [score - personal_model.predict([[i]])[0] for i, score in
    #                           enumerate(semester_average_scores)]
    #
    #     # 建立班级线性回归模型并计算每个学期的残差
    #     class_residuals = {}
    #     for class_name in set(classes):
    #         class_scores = [class_average_scores[i] for i in range(len(semesters)) if classes[i] == class_name]
    #         class_model = LinearRegression()
    #         class_model.fit([[i] for i in range(len(class_scores))], class_scores)
    #         class_residuals[class_name] = [score - class_model.predict([[i]])[0] for i, score in
    #                                        enumerate(class_scores)]
    #
    #     # 建立专业线性回归模型并计算每个学期的残差
    #     major_residuals = {}
    #     for major in set(majors):
    #         major_scores = [major_average_scores[i] for i in range(len(semesters)) if majors[i] == major]
    #         major_model = LinearRegression()
    #         major_model.fit([[i] for i in range(len(major_scores))], major_scores)
    #         major_residuals[major] = [score - major_model.predict([[i]])[0] for i, score in enumerate(major_scores)]
    #
    #     # 构建雷达图和折线图数据
    #     radar_data = []
    #     line_data = []
    #     radar_data = []
    #     for i in range(len(semesters)):
    #         radar_data.append({
    #             "subject": "个人增值",
    #             "score": personal_residuals[i],
    #             "type": semesters[i]
    #         })
    #         radar_data.append({
    #             "subject": "班级平均增值",
    #             "score": class_residuals[classes[i]][i] if classes[i] in class_residuals else 0,
    #             "type": semesters[i]
    #         })
    #         radar_data.append({
    #             "subject": "专业平均增值",
    #             "score": major_residuals[majors[i]][i] if majors[i] in major_residuals else 0,
    #             "type": semesters[i]
    #         })
    #         line_data.append({
    #             "subject": semesters[i],
    #             "学期平均分": semester_average_scores[i],
    #             "个人增值": personal_residuals[i],
    #             "班级平均增值": class_residuals[classes[i]][i] if classes[i] in class_residuals else 0,
    #             "专业平均增值": major_residuals[majors[i]][i] if majors[i] in major_residuals else 0
    #         })
    #     # print(radar_data)
    #     return {"radarData": radar_data, "lineData": line_data}
    @staticmethod
    def get_self_home_data_value_added(student_id):
        # SQL查询，确保选出所有选修了四门特定课程的学生及其成绩，同时避免重复记录
        sql_query = text("""
            SELECT 
                e.学号,
                e.课程名称,
                AVG(
                    CASE 
                        WHEN e.成绩 = '优秀' THEN 95
                        WHEN e.成绩 = '良好' THEN 85
                        WHEN e.成绩 = '中等' THEN 75
                        WHEN e.成绩 = '及格' THEN 60
                        WHEN e.成绩 = '通过' THEN 100
                        WHEN e.成绩 = '不及格' THEN 0
                        ELSE CAST(成绩 AS DECIMAL)
                    END
                ) AS 成绩
            FROM 
                enrollment e
            WHERE 
                e.课程名称 IN ('程序设计基础A', '高级程序设计A', '数据结构与算法', '操作系统')
            GROUP BY 
                e.学号, e.课程名称
        """)
        db_results = db.session.execute(sql_query).fetchall()

        # 提取数据
        data = []
        for row in db_results:
            sid, course, score = row
            if score is not None:
                data.append((sid, course, float(score)))

        # 准备训练数据
        student_ids, courses, scores = zip(*data)
        courses = np.array(courses).reshape(-1, 1)
        scores = np.array(scores)

        # One-hot encode the course names
        encoder = OneHotEncoder()
        X = encoder.fit_transform(courses).toarray()

        # 使用GBDT进行回归分析
        def fit_gbdt_model(X, y):
            model = GradientBoostingRegressor()
            model.fit(X, y)
            return model

        # 建立GBDT模型
        model = fit_gbdt_model(X, scores)

        # 准备目标学生的数据
        target_courses = ['程序设计基础A', '高级程序设计A', '数据结构与算法', '操作系统']
        target_X = encoder.transform(np.array(target_courses).reshape(-1, 1)).toarray()

        # 预测值
        predicted_scores = model.predict(target_X)

        # 获取目标学生实际成绩
        target_actual_scores = {course: None for course in target_courses}
        for sid, course, score in data:
            if sid == student_id:
                target_actual_scores[course] = score

        # 检查目标学生是否选修了所有课程
        if None in target_actual_scores.values():
            raise ValueError("目标学生未选修所有指定课程或未找到目标学生")

        # 计算残差
        residuals = {course: target_actual_scores[course] - predicted for course, predicted in
                     zip(target_courses, predicted_scores)}


        # 准备目标学生的数据（线性回归模型）
        target_data = [(course, score) for course, score in target_actual_scores.items()]
        target_courses_linear, target_scores_linear = zip(*target_data)

        # 将目标学生的课程名称转换为数值标签
        label_encoder_linear = LabelEncoder()
        target_courses_linear_encoded = label_encoder_linear.fit_transform(target_courses_linear).reshape(-1, 1)

        # 对目标学生的成绩建立线性回归模型
        model_linear = LinearRegression()
        model_linear.fit(target_courses_linear_encoded, target_scores_linear)

        # 预测目标学生的成绩（线性回归模型）
        target_courses_linear_encoded = label_encoder_linear.transform(target_courses_linear).reshape(-1, 1)
        predicted_scores_linear = model_linear.predict(target_courses_linear_encoded)

        # 计算残差（线性回归模型）
        residuals_linear = {course: actual - predicted for course, actual, predicted in
                            zip(target_courses_linear, target_scores_linear, predicted_scores_linear)}

        # 构建返回数据
        radar_data = []
        line_data = []

        for course, predicted in zip(target_courses, predicted_scores):
            radar_data.append({
                "subject": course,
                "score": target_actual_scores[course],
                "type": "实际"
            })
            radar_data.append({
                "subject": course,
                "score": predicted,
                "type": "预测"
            })
            radar_data.append({
                "subject": course,
                "score": residuals[course],
                "type": "增值"
            })

        for course, actual, predicted in zip(target_courses_linear, target_scores_linear, predicted_scores_linear):
            line_data.append({
                "subject": course,
                "实际成绩": actual,
                "个人预测成绩": predicted,
                "增值": residuals_linear[course]
            })

        return {"radarData": radar_data, "lineData": line_data, "residuals": residuals,
                "residuals_linear": residuals_linear}

    @staticmethod
    def get_bubble_data(class_id, course_name, chapter):
        print(class_id, course_name, chapter)
        class_id = '%' + class_id + '%'
        course_name = '%' + course_name + '%'
        chapter = '%' + chapter + '%'
        # 已经建立了数据库连接,并且可以使用db.session执行SQL查询
        sql_query = text("""
                SELECT
                    kp.知识点名称,
                    scores.分数,
                    COUNT(scores.学号) AS 人数
                FROM
                    knowledge_point kp
                    LEFT JOIN (
                        SELECT
                            sk.知识点id,
                            sk.分数,
                            s.学号
                        FROM
                            student_knowledge sk
                            JOIN student s ON sk.学号 = s.学号
                        WHERE
                            s.班级 LIKE :class_id AND
                            sk.分数 != -1
                    ) scores ON kp.知识点id = scores.知识点id
                WHERE
                    kp.所属课程 LIKE :course_name AND
                    kp.章节 LIKE :chapter
                GROUP BY
                    kp.知识点名称, scores.分数
            """)

        db_results = db.session.execute(sql_query, {
            'class_id': str(class_id),
            'course_name': str(course_name),
            'chapter': str(chapter)
        })

        result = []
        for knowledge_point, score, count in db_results:
            if score is not None:
                if (len(knowledge_point) > 8):
                    knowledge_point = knowledge_point[:4] + "..." + knowledge_point[-4:]
                result.append({
                    "分数": float(score),
                    "人数": count,
                    "知识点": knowledge_point
                })
        # print(result)

        return result
